import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import load_workbook

import os
print(os.getcwd()) 

#fonction pour afficher l'inventoire
def view_sheet(filename, sheet_name):
    wb = load_workbook(filename=filename)
    ws = wb[sheet_name]

    if ws.max_row == 0 or ws.max_column == 0:
        print("L'inventoire est vide.")
    else:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                print(cell.value, end='\t')
            print()  
#fonction pour suppeimer un élément dans un document excel
def delete_row_by_value(filename, sheet_name, search_value, column_index):
    wb = load_workbook(filename=filename)
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=column_index, max_col=column_index):
        for cell in row:
            if cell.value == search_value:
                ws.delete_rows(cell.row)
                print(search_value + ' a été supprimé de l"inventoire.')
                break

    wb.save(filename)

#fonction pour chercher les coordonnées d'un element en utilisant son nom

def find_coordinates(filename, sheet_name, search_value):
    wb = load_workbook(filename=filename)
    ws = wb[sheet_name]

    coordinates = None

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value == search_value:
                coordinates = (row, col)
                return coordinates

    return coordinates

#fonction pour modifier la valeur d'une cellule en utilisant ses coordonnées
def modify_cell_value(filename, sheet_name, coordinates, new_value):
    wb = load_workbook(filename=filename)
    ws = wb[sheet_name]

    if coordinates is not None:
        row, col = coordinates
        ws.cell(row=row, column=col, value=new_value)

    wb.save(filename)

#fonction pour entrer un valeur (puis le modifier en str) dans les marges

def simple_input_window(prompt):
    user_input = simpledialog.askstring("Entrée", prompt)
    return user_input

#fonction qui force l'utilisateur a choisir un des choix suggérés

def simple_choice_window(prompt, choices):
    user_choice = simpledialog.askstring("Entrée", f"{prompt}\nChoix: {', '.join(choices)}")
    while user_choice not in choices:
        user_choice = simpledialog.askstring("Entrée", f"Choix Invalide. {prompt}\nChoix: {', '.join(choices)}")
    return user_choice

#fonction pour ajouter un élément a l'inventoire

def add_item():
    name = simple_input_window('Entrez le nom du produit:')
    price = float(simple_input_window('Entrez le prix du produit par unité:'))
    category = simple_choice_window('Entrez la catégorie du produit :', ['Épicerie', 'Beauté et Personnel', 'Santé'])
    quantity = float(simple_input_window('Entrez la quantité du produit:'))

    new_row = (name, category, quantity, price)
    ws = wb.active
    ws.append(new_row)
    print(name + ' a été ajouté a l"inventoire.')
    wb.save('testp.xlsx')

#fonction pour supprimer un élément a l'inventoire

def del_item():
    search_value = simple_input_window('Saisissez le nom du produit à supprimer:')
    delete_row_by_value('testp.xlsx', 'Sheet1', search_value, 1)

#fonction pour mis a jour un élement de l'inventoire

def update_item():
    name = simple_input_window('Saisissez le nom du produit à mettre à jour:')
    search_value = name
    coordinates = find_coordinates('testp.xlsx', 'Sheet1', search_value)

    if coordinates is not None:
        choices = ['Catégorie', 'Prix', 'Quantité']
        selected_choice = simple_choice_window('Selectionner la proprieté a modifier', choices)

        if selected_choice == 'Catégorie':
            new_category = simple_choice_window('Entrer la nouvelle categorie:', ['Épicerie', 'Beauté et Personnel', 'Santé'])
            modify_cell_value('testp.xlsx', 'Sheet1', (coordinates[0], coordinates[1] + 1), new_category)
            print('Catégorie a été modifié.')

        elif selected_choice == 'Prix':
            new_price = float(simple_input_window('Enrer le nouveau prix:'))
            modify_cell_value('testp.xlsx', 'Sheet1', (coordinates[0], coordinates[1] + 3), new_price)
            print('Prix a été modifié.')

        elif selected_choice == 'Quantité':
            new_quantity = float(simple_input_window('Entrez la nouvelle quantité:'))
            modify_cell_value('testp.xlsx', 'Sheet1', (coordinates[0], coordinates[1] + 2), new_quantity)
            print(' prix a été modifié.')

    else:
        print(name + ' n"existe pas dans l"inventaire. Pensez à l"ajouter.')

#fonction pour faire des transaction

def perform_transaction():
    search_value = simple_input_window('Saisissez le nom du produit pour la transaction:')
    tr_num = int(simple_input_window('Saisissez le nombre de transactions concernées par unité du produit:'))

    item_coordinates = find_coordinates('testp.xlsx', 'Sheet1', search_value)
    if item_coordinates is not None:
        row, column = item_coordinates
        column += 3
        cell_value = ws.cell(row=row, column=column).value
        new_tr_value = cell_value * tr_num

        trcell_value = ws.cell(row=row, column=column + 1).value
        new_value_tr = trcell_value + new_tr_value

        coordinates_tr = (row, column)
        modify_cell_value('testp.xlsx', 'Sheet1', coordinates_tr, new_value_tr)

        print(f'Nouvelles transactions pour {search_value} sont {new_tr_value}')
        print(f'Nouveau total de transactions pour {search_value} est {new_value_tr}')

    else:
        print(f'{search_value} n"existe pas dans l"inventaire.')


#fonction qui affiche l'inventoire complète

def view_inventory():
    view_sheet('testp.xlsx', 'Sheet1')

#creation du fenètre principale
root = tk.Tk()
root.title('Gestion de l"inventaire')

# Creation des bouttons

add_button = tk.Button(root, text='Ajouter un nouvel article', command=add_item)
add_button.pack(pady=10)

delete_button = tk.Button(root, text='Effacer l"article', command=del_item)
delete_button.pack(pady=10)

update_button = tk.Button(root, text='Mettre à jour l"article', command=update_item)
update_button.pack(pady=10)

transaction_button = tk.Button(root, text='Effectuer une transaction', command=perform_transaction)
transaction_button.pack(pady=10)

view_button = tk.Button(root, text='Inspecter l"inventoire', command=view_inventory)
view_button.pack(pady=10)

exit_button = tk.Button(root, text='Sortire', command=root.destroy)
exit_button.pack(pady=10)

wb = load_workbook('testp.xlsx')
ws = wb.active

# Début du programme principale
root.mainloop()
